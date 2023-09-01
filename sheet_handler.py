from openpyxl import Workbook, load_workbook
import os.path

# TO DO: internally rename to sheet_edit


# Below checks if the file exists.
# If it does, then using openpyxl, we load the Workbook that we want to work with.
# Otherwise:
# - Set the name of the file we want it to be saved as
# - Open a new workbook
# - Insert the Card #, Card Name, and Quantity fields on the top row
# - Save the newly created workbook with the file name created earlier.

def createWorkbook():
    if (os.path.exists('./Digimon Card Catalog.xlsx')):
        workbook = load_workbook(filename="Digimon Card Catalog.xlsx")
        print("Digimon Card Catalog.xlsx loaded!")
    else:
        filename = "Digimon Card Catalog.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        
        sheet['A1'] = "Card #"
        sheet['B1'] = "Card Name"
        sheet['C1'] = "Quantity"

        sheet.title("BT1")
        
        workbook.save(filename=filename)
        print("Digimon Card Catalog.xlsx created!")
    return workbook

def saveWorkbook(workbook):
    workbook.save('Digimon Card Catalog.xlsx')
    print("Save completed!")

def addCards(workbook):
    toContinue = "Y"

    print("Welcome to the Add/Remove menu!\nTo add cards, enter a positive number at quantity.\nTo remove cards, enter a negative number at quantity.")

    # This while loop controls how many times someone goes through this list.
    
    while(toContinue == "Y"): 
        cardNum = str(input("\nCard Number: ")).upper() # Cast to string as Digimon card numbers are formatted as BT1-001, BT2-001, etc.
        try:
            if cardNum[3] == "-" or cardNum[3] == "=": # This catches all card numbers from BT1 - BT9.
                setFrom = cardNum[0:3]
            elif cardNum[0] == "P" and cardNum[1] == "-": # This catches Promo cards, which is formatted as P-001, P-002, etc.
                setFrom = "PROMOS"
            else:
                setFrom = cardNum[0:4] # This catches the rest, i.e. BT10, BT11, etc.
        except:
            print("Entry too short; please try again.") # Catches if someone enters something the length of 1 or 2.
            continue
        cardName = str(input("Card Name: "))
        quantity = int(input("# of Cards: "))

        # Below checks if setFrom (BT1, BT2, etc.) already exists in the loaded workbook.
        # If it isn't, then it creates a new sheet using the new set's booster number.
        # It then fills the cells with the default header rows.
        # If it does already exist, it makes that sheet the active sheet.
        
        if setFrom not in str(workbook.sheetnames):
            
            workbook.create_sheet(setFrom)
            print("\n" + setFrom + " sheet created!")
            workbook.active = workbook[setFrom]
            sheet = workbook.active
            
            sheet['A1'] = "Card #"
            sheet['B1'] = "Card Name"
            sheet['C1'] = "Quantity"
            
        else:
            workbook.active = workbook[setFrom]
            sheet = workbook.active
            print("\n" + setFrom + " sheet loaded!")

        # Giving the user a chance to error correct. 
        print("\nYou entered:\n  Card #: " + str(cardNum) + "\n  Card Name: " + str(cardName) + "\n  Quantity: " + str(quantity))
        correct = str(input("\nIf this is correct, enter Y; otherwise, enter N: ")).upper()

        while correct not in ["Y", "N"]:
            correct = str(input("\nPlease enter Y or N: ")).upper()

        # Below is how the script finds whether there is already an existing record of the card.
        # We set a "found" variable to 0 and a "max_row" variable to sheet.max_row+1 so we don't have to constantly call it.
        if correct == "Y":
            found = 0
            max_row = sheet.max_row+1
            for i in range(2, max_row):
                if str(sheet.cell(i, 1).value) == cardNum: # Check if the value of the cell is the same as the input...
                    newQuant = int(sheet.cell(i, 3).value) + quantity # If it is, then add the value of what's in cell 3 with the quantity input above...
                    sheet.cell(i, 3, value=newQuant) # Then update the cell with the new quantity.
                    found = 1 # Set found to 1...
                    break # So when the loop breaks, it avoids the if statement.
                
            # If the for loop fails, then go to max_row (the earliest row of cells) and input the new data.
            if found == 0:
                sheet.cell(max_row, 1, value=cardNum)
                sheet.cell(max_row, 2, value=cardName)
                sheet.cell(max_row, 3, value=quantity)
            
            saveWorkbook(workbook)
        else:
            print("\nStarting over...")
            continue

        # If the user wants to continue adding or removing cards, they enter "Y"; otherwise, they enter "N".
        toContinue =  str(input("\nContinue? Y/N: ")).upper()
        while toContinue not in ["Y", "N"]:
            toContinue =  str(input("Please enter Y/N: ")).upper()



# TO DO: Make this its own file with it's own logic.

def sortCards(workbook):
    pass
