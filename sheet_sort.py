from openpyxl import Workbook, load_workbook
import os.path

# TO DO: like. all of this.

def onStart():
    if (os.path.exists('./Digimon Card Catalog.xlsx')):
        workbook = load_workbook(filename="Digimon Card Catalog.xlsx")
        print("Digimon Card Catalog.xlsx loaded!")
    else:
        filename = "Digimon Card Catalog.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        workbook.save(filename=filename)
        print("Digimon Card Catalog.xlsx created!")
